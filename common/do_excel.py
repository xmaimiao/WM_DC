from openpyxl import load_workbook
class Case:
    def __init__(self):
        '''測試用例類，每個測試用例，實際上就是它的一個實例'''
        self.id = None
        self.username = None

class Do_Excel:
    def __init__(self,path,sheet_name):
        self.path = path
        self.sheet_name = sheet_name
        self.wb = load_workbook(path)
        self.sheet = self.wb[sheet_name]



    def read_excel(self):
        self.wb = load_workbook(self.path)
        self.sheet = self.wb[self.sheet_name]
        cases=[]
        # case = Case()    #注意：在這裡實例化只會取到最後一行數據
        for row in range(2,self.sheet.max_row+1): #注意這裏要從2開始
            case = Case()
            case.id = self.sheet.cell(row, 1).value
            case.username = self.sheet.cell(row, 2).value
            cases.append(case)
        return cases

    def write_excel(self,row,column,actual):
        self.sheet.cell(row,column).value = actual
        self.wb.save(self.path)
        self.wb.close()



